"""Professional Excel exports based on Pandas and OpenPyXL."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from analytics.exceptions import ExportError
from analytics.models import ExportResult, KPIResult
from analytics.utils import ensure_directory, require_dependency


@dataclass(frozen=True)
class ExcelSheet:
    """One worksheet to be written to an Excel workbook."""

    name: str
    dataframe: Any


class ExcelExporter:
    """Exports tabular analytics data to formatted Excel workbooks."""

    def export(
        self,
        path: str | Path,
        sheets: list[ExcelSheet],
        kpis: list[KPIResult] | None = None,
    ) -> ExportResult:
        """Write an Excel workbook with one or more formatted sheets."""

        require_dependency("pandas", "pandas")
        require_dependency("openpyxl", "openpyxl")
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        target = Path(path)
        ensure_directory(target.parent)
        if not sheets:
            raise ExportError("Nenhuma aba informada para exportação Excel.")

        try:
            with _excel_writer(target) as writer:
                if kpis:
                    _write_kpi_sheet(writer, kpis)
                for sheet in sheets:
                    sheet.dataframe.to_excel(writer, sheet_name=_sheet_name(sheet.name), index=False)

            from openpyxl import load_workbook

            workbook = load_workbook(target)
            for worksheet in workbook.worksheets:
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4F1B5C")
                for column_cells in worksheet.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column_cells)
                    letter = get_column_letter(column_cells[0].column)
                    worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)
            workbook.save(target)
            return ExportResult(path=target, format="xlsx")
        except Exception as exc:
            raise ExportError(str(exc)) from exc


def export_dataframe(path: str | Path, dataframe: Any, sheet_name: str = "Dados") -> ExportResult:
    """Export a single DataFrame to Excel."""

    return ExcelExporter().export(path, [ExcelSheet(sheet_name, dataframe)])


def _excel_writer(path: Path) -> Any:
    require_dependency("pandas", "pandas")
    import pandas as pd

    return pd.ExcelWriter(path, engine="openpyxl")


def _write_kpi_sheet(writer: Any, kpis: list[KPIResult]) -> None:
    require_dependency("pandas", "pandas")
    import pandas as pd

    rows = [
        {
            "Indicador": kpi.label or kpi.name,
            "Valor": kpi.value,
            "Unidade": kpi.unit or "",
        }
        for kpi in kpis
    ]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Resumo", index=False)


def _sheet_name(name: str) -> str:
    return (name or "Dados")[:31]
